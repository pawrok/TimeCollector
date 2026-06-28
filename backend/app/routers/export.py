import csv
import io
from datetime import date
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlmodel import Session

from ..database import get_session
from .. import crud

router = APIRouter(prefix="/export", tags=["export"])


def _build_table(db: Session):
    trackers = crud.get_trackers(db)
    rows = crud.get_daily_stats(db)
    data: dict[int, dict[str, float]] = {t.id: {} for t in trackers}
    all_dates: set[str] = set()
    for r in rows:
        day, tid, _, _, secs = r
        if tid in data:
            data[tid][day] = secs
            all_dates.add(day)
    return trackers, sorted(all_dates), data


@router.get("/excel")
def export_excel(db: Session = Depends(get_session)):
    from openpyxl import Workbook

    trackers, sorted_dates, data = _build_table(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Data"
    ws.cell(row=1, column=1, value="Date")
    for col, t in enumerate(trackers, start=2):
        ws.cell(row=1, column=col, value=t.name)
    for ri, day in enumerate(sorted_dates, start=2):
        ws.cell(row=ri, column=1, value=day)
        for col, t in enumerate(trackers, start=2):
            ws.cell(row=ri, column=col, value=round(data[t.id].get(day, 0) / 3600, 4))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"timecollector_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/csv")
def export_csv(db: Session = Depends(get_session)):
    trackers, sorted_dates, data = _build_table(db)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date"] + [t.name for t in trackers])
    for day in sorted_dates:
        writer.writerow(
            [day] + [round(data[t.id].get(day, 0) / 3600, 4) for t in trackers]
        )
    buf.seek(0)
    filename = f"timecollector_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
