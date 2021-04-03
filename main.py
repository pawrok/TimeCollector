from datetime import datetime, timedelta, date
from functools import partial
from tinydb import TinyDB, Query
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from kivy.config import Config
Config.set('graphics', 'width', '300')
Config.set('graphics', 'height', '650')

from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.dropdown import DropDown
from kivy.app import App
from kivy.lang import Builder
from kivy.properties import StringProperty, NumericProperty, DictProperty
from kivy.clock import Clock
from kivy.core.window import Window

import matplotlib.pyplot as plt
import matplotlib as mpl
import kivy

class RootLayout(BoxLayout):
    pass


class CollectorTextInput(TextInput):
    def __init__(self, **kwargs):
        super(CollectorTextInput, self).__init__(**kwargs)
        Window.bind(on_key_down=self._on_keyboard_down)
        Clock.schedule_once(self.collector_focus, 1)

    def collector_focus(self, *args):
        self.focus = True

    def _on_keyboard_down(self, instance, keyboard, keycode, text, modifiers):
        if self.focus and keycode == 40:  # Enter
            App.get_running_app().create_new_tracker(self.text)


class RenamePopup(Popup):
    tr_id = NumericProperty(0)

    def __init__(self, ID, **kwargs):
        super(RenamePopup, self).__init__(**kwargs)
        self.tr_id = ID
        Window.bind(on_key_down=self._on_keyboard_down)

    def _on_keyboard_down(self, instance, keyboard, keycode, text, modifiers):
        if self.rename_input.focus and keycode == 40:  # Enter
            App.get_running_app().rename_tracker(self.tr_id,
                                                 self.rename_input.text)
            self.dismiss()


class TrackerContainer(BoxLayout):
    ID = NumericProperty(0)
    timer_on = NumericProperty(0)
    name = StringProperty('')
    timer_event = None

    # time saved in db needed to load tracker with last time
    total_duration = NumericProperty(0.001)
    # current time based on current time and tracker start time diff
    current_duration = NumericProperty(0.001)
    # time stored after refresh, used to draw time label
    stored_time = NumericProperty(0.001)


class TimeTracker(App):
    def build(self):
        self.title = 'Time Collector'
        self.icon = 'icons/hourglass.png'
        self.trackers_indices = {}
        self.db = TinyDB('db.json')

        Builder.load_file('root.kv')
        self.root = RootLayout()
        self.load_trackers()
        self.reset_stats_at_new_day()
        return self.root

    def match_ids_to_indices(self):
        self.trackers_indices = {}
        for index in range(len(self.root.ids['box'].children)):
            id = self.root.ids['box'].children[index].ID
            self.trackers_indices[id] = index

    def start_timer(self, timer_id):
        self.stop_all_timers()
        now = datetime.now()
        event = Clock.schedule_interval(
            partial(self.update_label, start_time=now, timer_id=timer_id),
            0.06)

        timer_index = self.trackers_indices.get(timer_id)
        self.root.ids['box'].children[timer_index].timer_on = 1
        self.root.ids['box'].children[timer_index].ids['img'].source \
            = 'icons/pause.png'

        return event

    def start_stop_timer(self, timer_id):
        timer_index = self.trackers_indices.get(timer_id)
        tracker = self.root.ids['box'].children[timer_index]
        if not tracker.timer_on:
            tracker.timer_event = self.start_timer(timer_id)
        else:
            self.stop_timer(timer_id)

    def stop_timer(self, timer_id):
        timer_index = self.trackers_indices.get(timer_id)
        tracker = self.root.ids['box'].children[timer_index]
        tracker.timer_on = 0
        tracker.ids['img'].source = 'icons/play.png'
        tracker.timer_event.cancel()
        tracker.total_duration += tracker.current_duration
        tracker.stored_time += tracker.current_duration

        self.save_timer(timer_id, timer_index)

    def stop_all_timers(self):
        for id, index in self.trackers_indices.items():
            if self.root.ids['box'].children[index].timer_on == 1:
                self.stop_timer(id)

    def save_timer(self, tracker_id, timer_index):
        tracker = self.root.ids['box'].children[timer_index]
        tracker_data = self.db.search(Query().tracker_id == tracker_id)
        past_data = tracker_data[0]['past_data']

        today = date.today().strftime("%d.%m.%Y")

        past_data[today] += tracker.total_duration - past_data[today]

        self.db.update(
            {'total_time': tracker.total_duration, 'stored_time': tracker.stored_time},
            Query().tracker_id == tracker_id)

    def update_label(self, dt, start_time, timer_id):
        timer_index = self.trackers_indices.get(timer_id)
        tracker = self.root.ids['box'].children[timer_index]

        diff = datetime.now() - start_time
        tracker.current_duration = diff.total_seconds()

        display_time = tracker.current_duration + tracker.stored_time
        t = str(timedelta(seconds=display_time)).split('.')
        tracker.ids['time'].text = t[0] + '.' + t[1][0:2]

    def refresh_tracker_time(self, timer_id):
        timer_index = self.trackers_indices.get(timer_id)
        tracker = self.root.ids['box'].children[timer_index]

        if tracker.timer_on == 1:
            self.stop_timer(timer_id)
        else:
            tracker.stored_time = 0.001

            zero_time = str(timedelta(seconds=0.001)).split('.')
            tracker.ids['time'].text = zero_time[0] + '.' + zero_time[1][0:2]

            self.save_timer(timer_id, timer_index)

    def create_new_tracker(
            self, new_name, id=-1, total_time=0.001, stored_time=0.001):
        if id == -1:
            if self.db.all():
                max_id = sorted(self.db.all(), key=lambda k: k['tracker_id'])
                max_id = max_id[-1]['tracker_id']
            else:
                max_id = -1
            new_id = max_id + 1
            self.db.insert({
                    'tracker_name': new_name, 'tracker_id': new_id,
                    'total_time': total_time, 'stored_time': stored_time,
                    'past_data': {date.today().strftime("%d.%m.%Y"): 0}
                })
        else:
            new_id = id

        self.root.ids['box'].add_widget(
            TrackerContainer(ID=new_id, name=new_name,
                             total_duration=total_time,
                             stored_time=stored_time))
        self.root.ids['box'].height += \
            TrackerContainer.height.defaultvalue * 1.2   # 1.2 == padding

        self.match_ids_to_indices()

        timer_index = self.trackers_indices.get(new_id)
        t = str(timedelta(seconds=stored_time)).split('.')
        self.root.ids['box'].children[timer_index].ids['time'].text \
            = t[0] + '.' + t[1][0:2]

    def delete_tracker(self, timer_id):
        timer_index = self.trackers_indices.get(timer_id)
        tracker = self.root.ids['box'].children[timer_index]

        if tracker.timer_on == 1:
            self.stop_timer(timer_id)

        self.root.ids['box'].remove_widget(tracker)
        self.match_ids_to_indices()
        
        self.root.ids['box'].height -= TrackerContainer.height.defaultvalue
        self.db.remove(Query().tracker_id == timer_id)

    def rename_tracker(self, timer_id, new_name):
        timer_index = self.trackers_indices.get(timer_id)
        tracker = self.root.ids['box'].children[timer_index]
        tracker.name = new_name
        self.db.update(
            {'tracker_name': new_name}, Query().tracker_id == timer_id)

    def load_trackers(self):
        for item in self.db:
            self.create_new_tracker(
                item['tracker_name'], item['tracker_id'],
                item['total_time'], item['stored_time'])

    def on_stop(self):
        " saves all timer's time when exiting the app "
        self.stop_all_timers()

    def reset_stats_at_new_day(self):
        today = date.today().strftime("%d.%m.%Y")
        trackers = self.root.ids['box'].children

        for index in range(len(trackers)):
            tracker_id = trackers[index].ID

            tracker_data = self.db.search(Query().tracker_id == tracker_id)
            past_data = tracker_data[0]['past_data']

            if today not in past_data:
                past_data[today] = 0
                self.db.update({'total_time': 0.001, 'stored_time': 0.001},
                               Query().tracker_id == id)

    def plot_past_data(self):
        all_data = self.db.all()
        # plt.xkcd()
        mpl.rcParams.update({'font.size': 17})
        plt.style.use('dark_background')

        ' linear, every tracker '
        for item in all_data:
            plot_date = []
            plot_duration = []
            for key, value in item['past_data'].items():
                plot_date.append(int(key.split('.')[0]))
                plot_duration.append(value/3600)
            plt.plot(plot_date, plot_duration, 'wo')
            plt.plot(plot_date, plot_duration)

        plt.title('Time spent each day', fontsize=22)
        plt.xlabel('Day', fontsize=18)
        plt.ylabel('Time (h)', fontsize=18)
        plt.gcf().subplots_adjust(bottom=0.15)
        plt.gcf().subplots_adjust(left=0.15)
        plt.savefig("linear.png")

        ' pie '
        labels = []
        durations = []
        for item in all_data:
            labels.append(item['tracker_name'].replace(" ", "\n"))
            sum_h = 0
            for key, value in item['past_data'].items():
                sum_h += value
            durations.append(sum_h/3600)

        def func(pct, allvals):
            absolute = int(pct/100.*sum(allvals))
            if pct < 3:
                return ""
            else:
                return "{:.1f}%\n({:d} h)".format(pct, absolute)

        fig1, ax1 = plt.subplots()
        explode = len(all_data) * (0.05,)
        patches, texts, autotexts = ax1.pie(
            durations, labels=labels, autopct=lambda pct: func(pct, durations),
            startangle=90, pctdistance=0.72, explode=explode)

        for text in texts:
            text.set_color('white')
            text.set_fontsize(12)

        for autotext in autotexts:
            autotext.set_color('black')
            autotext.set_fontsize(12)

        # draw circle
        centre_circle = plt.Circle((0, 0), 0.5, fc='black')
        fig = plt.gcf()
        fig.gca().add_artist(centre_circle)

        plt.title('Total time comparison', fontsize=18)
        plt.savefig("pie.png", bbox_inches='tight', pad_inches=0)

    def export_to_excel(self):
        wb = Workbook()
        data = self.db.all()
        ws = wb.active      # get worksheet

        for col in range(1, len(data) + 1):
            ws.cell(column=col * 2 - 1, row=1,
                    value=data[col - 1]['tracker_name'])
            ws.cell(column=col * 2 - 1, row=2, value='date')
            ws.cell(column=col * 2, row=2, value='time (s)')

            past_data = data[col-1]['past_data']
            dates = list(past_data.keys())
            times = list(past_data.values())

            for i in range(len(past_data)):
                ws.cell(column=col * 2 - 1, row=i + 3, value=dates[i])
                ws.cell(column=col * 2, row=i + 3, value=times[i])

        wb.save("collector_data.xlsx")


if __name__ == '__main__':
    TimeTracker().run()
